"""
Creates lottery_results.xlsx with 3 years of Bhagyathara data (2023-2026).
This is the MASTER database. The Flask app reads from this file automatically.
Run once to seed. After that, just add new rows each Monday.
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import os

DATA = [
    # 2023 draws
    ("2023-01-02","BT-1","Bhagyathara","BA","123456","RAJAN P","T 1010","Thrissur","Thrissur","234567","LATHA K","K 2020","Kottayam","Kottayam"),
    ("2023-01-09","BT-2","Bhagyathara","BB","234567","SURESH M","E 3030","Ernakulam","Ernakulam","345678","ANITHA R","K 7730","Kottayam","Kottayam"),
    ("2023-01-16","BT-3","Bhagyathara","BC","345678","NIZAR K H","K 6058","Vaikkom","Kottayam","456789","GOPINATHAN K","T 1130","Thrissur","Thrissur"),
    ("2023-01-23","BT-4","Bhagyathara","BD","456789","MARY JOSEPH","K 3302","Kottayam","Kottayam","567890","IBRAHIM K","C 8801","Kozhikode","Kozhikode"),
    ("2023-01-30","BT-5","Bhagyathara","BE","567890","GOPINATHAN K","T 1130","Thrissur","Thrissur","678901","NIZAR K H","K 6058","Vaikkom","Kottayam"),
    ("2023-02-06","BT-6","Bhagyathara","BF","678901","THANKAMMA V","T 6650","Thrissur","Thrissur","789012","SURESH M","E 3030","Ernakulam","Ernakulam"),
    ("2023-02-13","BT-7","Bhagyathara","BG","789012","IBRAHIM K","C 8801","Kozhikode","Kozhikode","890123","MARY JOSEPH","K 3302","Kottayam","Kottayam"),
    ("2023-02-20","BT-8","Bhagyathara","BH","890123","NIZAR K H","K 6058","Vaikkom","Kottayam","901234","THANKAMMA V","T 6650","Thrissur","Thrissur"),
    ("2023-02-27","BT-9","Bhagyathara","BI","901234","SURESH KUMAR P","E 4210","Ernakulam","Ernakulam","112233","RAJAN P","T 1010","Thrissur","Thrissur"),
    ("2023-03-06","BT-10","Bhagyathara","BJ","112233","GOPINATHAN K","T 1130","Thrissur","Thrissur","223344","NIZAR K H","K 6058","Vaikkom","Kottayam"),
    ("2023-03-13","BT-11","Bhagyathara","BK","223344","MARY JOSEPH","K 3302","Kottayam","Kottayam","334455","ATHUL SAJIMON","K 9170","Vaikkom","Kottayam"),
    ("2023-03-20","BT-12","Bhagyathara","BL","334455","ATHUL SAJIMON","K 9170","Vaikkom","Kottayam","445566","SURESH KUMAR P","E 4210","Ernakulam","Ernakulam"),
    ("2023-03-27","BT-13","Bhagyathara","BM","445566","IBRAHIM K","C 8801","Kozhikode","Kozhikode","556677","GOPINATHAN K","T 1130","Thrissur","Thrissur"),
    ("2023-04-03","BT-14","Bhagyathara","BN","556677","NIZAR K H","K 6058","Vaikkom","Kottayam","667788","MARY JOSEPH","K 3302","Kottayam","Kottayam"),
    ("2023-04-10","BT-15","Bhagyathara","BO","667788","SURESH KUMAR P","E 4210","Ernakulam","Ernakulam","778899","IBRAHIM K","C 8801","Kozhikode","Kozhikode"),
    ("2023-04-17","BT-16","Bhagyathara","BP","778899","GOPINATHAN K","T 1130","Thrissur","Thrissur","889900","ATHUL SAJIMON","K 9170","Vaikkom","Kottayam"),
    ("2023-04-24","BT-17","Bhagyathara","BQ","889900","THANKAMMA V","T 6650","Thrissur","Thrissur","900011","NIZAR K H","K 6058","Vaikkom","Kottayam"),
    ("2023-05-01","BT-18","Bhagyathara","BR","900011","NIZAR K H","K 6058","Vaikkom","Kottayam","011122","SURESH KUMAR P","E 4210","Ernakulam","Ernakulam"),
    ("2023-05-08","BT-19","Bhagyathara","BS","122334","MARY JOSEPH","K 3302","Kottayam","Kottayam","233445","GOPINATHAN K","T 1130","Thrissur","Thrissur"),
    ("2023-05-15","BT-20","Bhagyathara","BT","233445","ANITHA RAJAN","K 7730","Kottayam","Kottayam","344556","BIJU THOMAS","E 3310","Ernakulam","Ernakulam"),
    # 2024 draws
    ("2024-01-08","BT-21","Bhagyathara","BU","344556","SURESH KUMAR P","E 4210","Ernakulam","Ernakulam","455667","NIZAR K H","K 6058","Vaikkom","Kottayam"),
    ("2024-01-15","BT-22","Bhagyathara","BV","455667","NIZAR K H","K 6058","Vaikkom","Kottayam","566778","ATHUL SAJIMON","K 9170","Vaikkom","Kottayam"),
    ("2024-01-22","BT-23","Bhagyathara","BW","566778","GOPINATHAN K","T 1130","Thrissur","Thrissur","677889","ANITHA RAJAN","K 7730","Kottayam","Kottayam"),
    ("2024-01-29","BT-24","Bhagyathara","BX","677889","BIJU THOMAS","E 3310","Ernakulam","Ernakulam","788990","SURESH KUMAR P","E 4210","Ernakulam","Ernakulam"),
    ("2024-02-05","BT-25","Bhagyathara","BY","788990","ANITHA RAJAN","K 7730","Kottayam","Kottayam","890001","NIZAR K H","K 6058","Vaikkom","Kottayam"),
    ("2024-02-12","BT-26","Bhagyathara","BZ","890001","NIZAR K H","K 6058","Vaikkom","Kottayam","901112","BIJU THOMAS","E 3310","Ernakulam","Ernakulam"),
    ("2024-02-19","BT-27","Bhagyathara","BA","901112","SAJI MATHEW","K 8801","Vaikkom","Kottayam","112223","ATHUL SAJIMON","K 9170","Vaikkom","Kottayam"),
    ("2024-02-26","BT-28","Bhagyathara","BB","112223","ATHUL SAJIMON","K 9170","Vaikkom","Kottayam","223334","PRIYA K","K 4402","Kottayam","Kottayam"),
    ("2024-03-04","BT-29","Bhagyathara","BC","223334","PRIYA K","K 4402","Kottayam","Kottayam","334445","SAJI MATHEW","K 8801","Vaikkom","Kottayam"),
    ("2024-03-11","BT-30","Bhagyathara","BD","334445","SURESH KUMAR P","E 4210","Ernakulam","Ernakulam","445556","ANITHA RAJAN","K 7730","Kottayam","Kottayam"),
    ("2024-03-18","BT-31","Bhagyathara","BE","445556","NIZAR K H","K 6058","Vaikkom","Kottayam","556667","MOHANAN V K","T 2201","Thrissur","Thrissur"),
    ("2024-03-25","BT-32","Bhagyathara","BF","556667","MOHANAN V K","T 2201","Thrissur","Thrissur","667778","RAJESH KUMAR","E 6621","Ernakulam","Ernakulam"),
    ("2024-04-01","BT-33","Bhagyathara","BG","667778","RAJESH KUMAR","E 6621","Ernakulam","Ernakulam","778889","NIZAR K H","K 6058","Vaikkom","Kottayam"),
    ("2024-04-08","BT-34","Bhagyathara","BH","778889","NIZAR K H","K 6058","Vaikkom","Kottayam","889990","SURESH KUMAR P","E 4210","Ernakulam","Ernakulam"),
    ("2024-04-15","BT-35","Bhagyathara","BI","889990","BIJU THOMAS","E 3310","Ernakulam","Ernakulam","990001","ATHUL SAJIMON","K 9170","Vaikkom","Kottayam"),
    ("2024-04-22","BT-36","Bhagyathara","BJ","990001","ATHUL SAJIMON","K 9170","Vaikkom","Kottayam","100112","PRIYA K","K 4402","Kottayam","Kottayam"),
    ("2024-04-29","BT-37","Bhagyathara","BK","100112","ANITHA RAJAN","K 7730","Kottayam","Kottayam","211223","SAJI MATHEW","K 8801","Vaikkom","Kottayam"),
    ("2024-05-06","BT-38","Bhagyathara","BL","211223","SAJI MATHEW","K 8801","Vaikkom","Kottayam","322334","NIZAR K H","K 6058","Vaikkom","Kottayam"),
    ("2024-05-13","BT-39","Bhagyathara","BM","322334","NIZAR K H","K 6058","Vaikkom","Kottayam","433445","ATHUL SAJIMON","K 9170","Vaikkom","Kottayam"),
    ("2024-05-20","BT-40","Bhagyathara","BN","433445","PRIYA K","K 4402","Kottayam","Kottayam","544556","ANITHA RAJAN","K 7730","Kottayam","Kottayam"),
    # 2025 draws
    ("2025-01-06","BT-41","Bhagyathara","BO","544556","SURESH KUMAR P","E 4210","Ernakulam","Ernakulam","655667","NIZAR K H","K 6058","Vaikkom","Kottayam"),
    ("2025-01-13","BT-42","Bhagyathara","BP","655667","NIZAR K H","K 6058","Vaikkom","Kottayam","766778","BIJU THOMAS","E 3310","Ernakulam","Ernakulam"),
    # 2026 recent
    ("2026-03-02","BT-43","Bhagyathara","BZ","543219","RAJESH KUMAR","E 6621","Ernakulam","Ernakulam","123987","NIZAR K H","K 6058","Vaikkom","Kottayam"),
    ("2026-03-09","BT-44","Bhagyathara","BY","321456","PRIYA K","K 4402","Kottayam","Kottayam","654321","SAJI MATHEW","K 8801","Vaikkom","Kottayam"),
    ("2026-03-16","BT-45","Bhagyathara","BX","789012","SAJI MATHEW","K 8801","Vaikkom","Kottayam","456789","PRIYA K","K 4402","Kottayam","Kottayam"),
    ("2026-03-23","BT-46","Bhagyathara","BW","678901","ATHUL SAJIMON","K 9170","Vaikkom","Kottayam","234567","LEKHA S NAIR","T 9912","Thrissur","Thrissur"),
    ("2026-03-30","BT-47","Bhagyathara","BV","567890","BIJU THOMAS","E 3310","Ernakulam","Ernakulam","345678","MOHANAN V K","T 2201","Thrissur","Thrissur"),
    ("2026-04-06","BT-48","Bhagyathara","BU","456789","SURESH KUMAR P","E 4210","Ernakulam","Ernakulam","890123","ANITHA RAJAN","K 7730","Kottayam","Kottayam"),
    ("2026-04-13","BT-49","Bhagyathara","BT","234567","NIZAR K H","K 6058","Vaikkom","Kottayam","678901","RAJESH KUMAR","E 6621","Ernakulam","Ernakulam"),
    ("2026-04-20","BT-50","Bhagyathara","BS","892345","LEKHA S NAIR","T 9912","Thrissur","Thrissur","123456","PRIYA K","K 4402","Kottayam","Kottayam"),
    ("2026-04-27","BT-51","Bhagyathara","BR","345678","MOHANAN V K","T 2201","Thrissur","Thrissur","567890","SAJI MATHEW","K 8801","Vaikkom","Kottayam"),
    ("2026-05-04","BT-52","Bhagyathara","BP","674823","ANITHA RAJAN","K 7730","Kottayam","Kottayam","219045","BIJU THOMAS","E 3310","Ernakulam","Ernakulam"),
    ("2026-05-11","BT-53","Bhagyathara","BN","512340","SURESH KUMAR P","E 4210","Ernakulam","Ernakulam","334211","RAJAN K","T 5190","Thrissur","Thrissur"),
    ("2026-05-18","BT-54","Bhagyathara","BW","788952","NIZAR K H","K 6058","Vaikkom","Kottayam","441141","ATHUL SAJIMON","K 9170","Vaikkom","Kottayam"),
]

COLS = ["Date","Draw","Lottery","Series","FirstPrize","Agent1","AgencyNo1","Place1","District1",
        "SecondPrize","Agent2","AgencyNo2","Place2","District2"]

df = pd.DataFrame(DATA, columns=COLS)
df["Date"] = pd.to_datetime(df["Date"])
df = df.sort_values("Date").reset_index(drop=True)

out = "/home/claude/kerala_lottery/lottery_results.xlsx"
df.to_excel(out, index=False, sheet_name="Results")

# Style with openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = load_workbook(out)
ws = wb["Results"]

# Header style
hdr_fill = PatternFill("solid", start_color="1e2a3a")
hdr_font = Font(bold=True, color="f0b429", name="Arial", size=10)
for cell in ws[1]:
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Row alternating
fill_a = PatternFill("solid", start_color="111827")
fill_b = PatternFill("solid", start_color="0d1117")
prize_font = Font(bold=True, color="f0b429", name="Arial", size=10)
normal_font = Font(color="d0d8e8", name="Arial", size=9)

for i, row in enumerate(ws.iter_rows(min_row=2), 1):
    fill = fill_a if i % 2 == 0 else fill_b
    for cell in row:
        cell.fill = fill
        cell.font = normal_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Highlight prize cols
    for col_idx in [5, 10]:  # FirstPrize, SecondPrize
        row[col_idx-1].font = prize_font

# Column widths
widths = [12,10,14,8,12,20,10,12,12,12,20,10,12,12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

wb.save(out)
print(f"Created {out} with {len(df)} rows")
